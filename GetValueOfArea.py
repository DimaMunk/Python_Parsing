import openpyxl
from openpyxl.cell import get_column_letter

def GetValueOfArea (sheet, column, row):
    idx = '{}{}'.format(get_column_letter(column), row)
    for range in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(range))
        for row in merged_cells:
            if idx in row:
                return sheet.cell(merged_cells[0][0]).value
    return sheet.cell(idx).value

